import os
import random
import re
import time 
import requests
import json
from csv import DictWriter 



from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from fake_useragent import UserAgent
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import Workbook, load_workbook


from bs4 import BeautifulSoup as bs
from bs4 import SoupStrainer as ss

driver_path = os.path.abspath(os.path.join(os.path.dirname(
    os.path.abspath(__file__)), 'driver'))

exel_path = os.path.abspath(os.path.join(os.path.dirname(
    os.path.abspath(__file__)), 'url_form.xlsx'))

urls_path = os.path.abspath(os.path.join(os.path.dirname(
    os.path.abspath(__file__)), 'urls.csv'))

options = Options()
ua = UserAgent(verify_ssl=False)

def grp(pat, txt):
    r = re.search(pat, txt)
    return r.group(0) if r else '&'

browsers = {
    'chrome': r'Chrome/[^ ]+',
    'safari': r'AppleWebKit/[^ ]+',
    'opera': r'Opera\s.+$',
    'firefox': r'Firefox/.+$',
    'internetexplorer': r'Trident/[^;]+',
}

browser = random.choice(ua.data_randomize)
userAgent = sorted(ua.data_browsers[browser], key=lambda a: grp(browsers[browser], a))[-1]
options.add_argument(f'user-agent={userAgent}')
options.add_argument('headless')


chop = webdriver.ChromeOptions()
chop.add_extension(f'{driver_path}/block.crx')

prefs = {"profile.managed_default_content_settings.images": 2}
chop.add_experimental_option("prefs", prefs)

field_names = ['url','state','price']


class Scraper :

    def __init__(self):
        self.driver = webdriver.Chrome(executable_path=f'{driver_path}/chromedriver',chrome_options = chop)
        self.wait =  WebDriverWait(self.driver, 10)

        # self.STATES = {
        #     'Florida':'5460aeac2e5b2','California':'5460aeaa10951','Colorado':'5460aea9a56f2','Alabama':'5460aeac1cd07','Alaska':'5460aea9a280a',
        #     'Arizona':'5460aeac241a8','Arkansas':'5460aea9b62b8','Connecticut':'5460aea9a7021','Delaware':'5460aeac2ddf1',
        #     'Georgia':'5460aea9ba5c1','Hawaii':'5460aeac2226a','Idaho':'5460aeac44938','Illinois':'5460aea9e8136','Indiana':'5460aeac4aae2',
        #     'Iowa':'5460aea9e847f','Kansas':'5460aea22af76','Kentucky':'5460aece68397','Louisiana':'5460aece67bc1','Maine':'5460aea22fd94',
        #     'Maryland':'5460aece6ca85','Massachusetts':'5460aece6bb84','Michigan':'5460aea23b15b','Minnesota':'5460aeaff0b61','Mississippi':'5460aeb000b77',
        #     'Missouri':'5460aea24d0b8','Montana':'5460aea2510f6','Nebraska':'5460aeaff3a38','Nevada':'5460aea258e1f','New-Hampshire':'5460aea24ae57',
        #     'New-Jersey':'5460aeb002e9e','New-Mexico':'5460aea260ede','New-York':'5460aeb030147','North-Carolina':'5460aea27a131','North-Dakota':'5460ae9c79c01',
        #     'Ohio':'5460aeb03ef8f','Oklahoma':'5460aea280e74','Oregon':'5460aeb03d7f9','Pennsylvania':'5460aeb044982','RhodeIsland':'5460aeb04939e',
        #     'South-Carolina':'5460aea281d95','South-Dakota':'5460aeb042602','Tennessee':'5460aea28cdf7','Texas':'5460aeb04ddd7','Utah':'5460aeb05b0a2','Vermont':'5460aea296a1a',
        #     'Virginia':'5460aeb05cfe8','Washington':'5c1cd29e065e3','West-Virginia':'5460aea2bc101','Wisconsin':'5460aeb061a1f','Wyoming':'5460aea2b1e26','Canada':'53be518a576fc'
        #     }
        self.STATES = {
            'Florida - Anna Maria Island':'5460aea9bc53f','Florida - Bradenton':'5460aeac2c62f','Florida - Holmes Beach':'5460aea9bcf6c','Florida - Fort Pierce':'5460aeb785c6f',
            'Florida - Sanford':'5460aeb785c6f','Florida - flagler county':'53e4e86935a1d','Florida - St Lucie county':'53e4e99811c37'
        }
        self.requests_session = requests.Session()
       
    def __open_browser(self,state):
        url = f'https://www.casamundo.com/search/{state}'
        self.driver.get(url)
        self.driver.set_window_size(500, 1000)
        self.driver.switch_to.window(self.driver.window_handles[0])

    def __scroll_down(self,current_scroll_position,speed=15,new_height=0):
        current_scroll_position, new_height= 0, 1
        while current_scroll_position <= new_height:
            current_scroll_position += speed
            self.__driver.execute_script("window.scrollTo(0, {});".format(current_scroll_position))
            new_height = self.__driver.execute_script("return document.body.scrollHeight")
    

    def __write_to_exel(self,filepath, dictionary):
        wb = Workbook(write_only=True)
        sheet = wb.create_sheet()

        headers = list(dictionary[0])
        sheet.append(headers)

        for x in dictionary:
            sheet.append(list(x.values()))

        wb.save(filepath)


    def __get_data(self,state_id,state):
        test = self.__open_browser(state_id)
        pos = 0
        n_height = 0

        try:
            while True:
                time.sleep(2)
                self.driver.execute_script(f"window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(2)
                btn = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,'div[class="df jcc filter-pager offers-footer pt16 mb32 pt8-xs ph8-xs tac posr clb"]')))
                btn.click()

        except:
            self.__scroll_down(current_scroll_position=0)
            html = bs(self.driver.page_source, "lxml")
            items = [item for item in html.find_all('div', attrs={'data-test' : 'grid-item'})]
            res = []
            for x in items :
                try :
                    obj = {}
                    unique_id = x.find('article' , {'class':'offer'}).get('id')

                    try:
                        price = x.find('div', attrs={'data-test' : 'price'}).find('div',attrs={'class':['fz32','lh32','fw500','c-gray-extra-dark']}).text
                    except:
                        pass

                    if unique_id:
                        obj['url'] = f'https://www.casamundo.com/rental/offer/{unique_id}'
                        obj['state'] = state
                        obj['price'] = price if price else '-'
                        res.append(obj)
                except:
                    pass
            return res

    def run(self):

        for key in self.STATES :
            res = self.__get_data(self.STATES[key],key)
            
            with open(urls_path, 'a') as f_object: 

                dictwriter_object = DictWriter(f_object, fieldnames=field_names) 

                for x in res:
                    dictwriter_object.writerow(x) 
                
                f_object.close()

        self.driver.quit()


if __name__ == '__main__':
    s = Scraper()
    s.run()
